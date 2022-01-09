'''
Name: Kumo Pascaline Myenneh
Email: kumo.pascaline@ictuniversity.edu.cm
ICTU Matricule: ICTU20201168
Contact: 661253212
Course: Programming in python
Course Instructor: Mr Fru Emmanuel
this is a python  program that access, edits and update emails in an excel file from
 @helpinghands.org to @helpinghands.cm
'''

import openpyxl as xl
import csv
#loads excel file
wb = xl.load_workbook('employeedata.xlsx')
#open workbook
sheet = wb.active # sheet 1 is the active cell thats y it is called in the workbook(wb)

old_domain = 'helpinghands.cm'
new_domain = 'handsinhands.org'

# ------- CREATING THE EMAILS USING THE USER'S NAME------------#

# ------- CREATING AND UPDATING A .xlsx file------------#
#modify the disired cell
for i in range(2,sheet.max_row + 1):

    # begins from two because we're working with the second row elements
    
    # reads the element of the cell e.g pascaline@helpinghands.cm changes it to @handsinhands.org. to it
    cell = sheet.cell(i,3)

    #----- replacing the old cell to store the new email -----#
    if old_domain in cell.value:
        #updating and replacing the old emails with the new ones
        updated_email=(cell.value).replace(old_domain,new_domain)# stores the new email in the 4th column(column D)

    #------ attributing the data to the very cell (column) having the old emails -----#
        sheet.cell(i,3).value = updated_email
    

 # ------- CREATING AND UPDATING A .csv FILE------------#

#reading the CSV file
text = open("employeedata.csv","r")
#join() method combines all contents of
#employeedata.csv and formed as a string
text = ''.join([i for i in text])

#search and replacee the contents
text = text.replace(old_domain, new_domain)

#updated_data.csv is the output file in write mode
x = open("updated_data.csv","w")

#all the replaced text is written in the update_data.csv file
x.writelines(text)
x.close()

# ----- creating a new document that holds the updated emails -----#
#i.e saving file as updated_data.xlsx
wb.save('updated_data.xlsx')



